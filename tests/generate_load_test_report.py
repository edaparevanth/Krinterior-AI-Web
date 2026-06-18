import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def main():
    json_path = "test_reports/load_test_results.json"
    output_path = "test_reports/Krinterior_AI_Load_Test_Report.xlsx"
    
    if not os.path.exists(json_path):
        print(f"Error: Load test results JSON not found at: {json_path}")
        print("Please run tests/load_test.py first to generate results.")
        return

    print(f"Reading load test results from {json_path}...")
    with open(json_path, "r") as f:
        data = json.load(f)
        
    summary = data["summary"]
    endpoints = data["endpoints"]
    
    print("Generating Load Test Excel Report...")
    wb = openpyxl.Workbook()
    
    # ------------------ SHEET 1: SUMMARY ------------------
    ws_summary = wb.active
    ws_summary.title = "Load Test Overview"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Color palette (Sleek Slate Blue / Ice Blue theme)
    PRIMARY_COLOR = "1F4E79"      # Steel Blue
    ACCENT_COLOR = "DDEBF7"       # Soft Ice Blue
    PASS_FILL_COLOR = "E2EFDA"    # Soft Sage Green
    PASS_TEXT_COLOR = "375623"    # Dark Sage Green
    FAIL_FILL_COLOR = "FCE4D6"    # Soft Peach
    FAIL_TEXT_COLOR = "C00000"    # Dark Red
    WHITE = "FFFFFF"
    
    # Fonts
    font_title = Font(name="Segoe UI", size=16, bold=True, color=WHITE)
    font_section = Font(name="Segoe UI", size=12, bold=True, color=PRIMARY_COLOR)
    font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)
    font_kpi_label = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_kpi_num = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    
    # Fills
    fill_primary = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_accent = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    fill_pass = PatternFill(start_color=PASS_FILL_COLOR, end_color=PASS_FILL_COLOR, fill_type="solid")
    fill_fail = PatternFill(start_color=FAIL_FILL_COLOR, end_color=FAIL_FILL_COLOR, fill_type="solid")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    thick_bottom_border = Border(
        bottom=Side(style='medium', color=PRIMARY_COLOR)
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='double', color=PRIMARY_COLOR)
    )

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # 1. Title Banner
    ws_summary.merge_cells("A1:H2")
    title_cell = ws_summary["A1"]
    title_cell.value = "KRINTERIOR AI - BACKEND LOAD TEST REPORT"
    title_cell.font = font_title
    title_cell.fill = fill_primary
    title_cell.alignment = align_center

    # 2. KPI Cards (Row 4 & 5)
    kpis = [
        {"col_start": 1, "col_end": 2, "label": "CONCURRENT VIRTUAL USERS", "val": summary["concurrent_users"], "fmt": "0"},
        {"col_start": 3, "col_end": 4, "label": "TEST DURATION (SECS)", "val": summary["test_duration_seconds"], "fmt": "0.0"},
        {"col_start": 5, "col_end": 6, "label": "TOTAL REQUESTS SENT", "val": summary["total_requests"], "fmt": "#,##0"},
        {"col_start": 7, "col_end": 8, "label": "SUCCESS RATE", "val": summary["success_rate"] / 100.0, "fmt": "0.0%"}
    ]

    for kpi in kpis:
        c1 = col_letter = get_column_letter(kpi["col_start"])
        c2 = col_letter = get_column_letter(kpi["col_end"])
        ws_summary.merge_cells(f"{c1}4:{c2}4")
        ws_summary.merge_cells(f"{c1}5:{c2}5")
        
        lbl_cell = ws_summary[f"{c1}4"]
        lbl_cell.value = kpi["label"]
        lbl_cell.font = font_kpi_label
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_accent
        
        val_cell = ws_summary[f"{c1}5"]
        val_cell.value = kpi["val"]
        val_cell.font = font_kpi_num
        val_cell.alignment = align_center
        val_cell.number_format = kpi["fmt"]
        
        # Apply borders to merged range
        for r in [4, 5]:
            for c in range(kpi["col_start"], kpi["col_end"] + 1):
                cell = ws_summary.cell(row=r, column=c)
                cell.border = thin_border

    # Additional KPI Row (Row 7 & 8)
    kpis2 = [
        {"col_start": 1, "col_end": 2, "label": "AVERAGE THROUGHPUT (RPS)", "val": summary["average_rps"], "fmt": "0.0"},
        {"col_start": 3, "col_end": 4, "label": "AVERAGE LATENCY (MS)", "val": summary["avg_response_time_ms"], "fmt": "0.0"},
        {"col_start": 5, "col_end": 6, "label": "95th PERCENTILE LATENCY", "val": summary["p95_response_time_ms"], "fmt": "0.0"},
        {"col_start": 7, "col_end": 8, "label": "MAX LATENCY (MS)", "val": summary["max_response_time_ms"], "fmt": "0.0"}
    ]

    for kpi in kpis2:
        c1 = col_letter = get_column_letter(kpi["col_start"])
        c2 = col_letter = get_column_letter(kpi["col_end"])
        ws_summary.merge_cells(f"{c1}7:{c2}7")
        ws_summary.merge_cells(f"{c1}8:{c2}8")
        
        lbl_cell = ws_summary[f"{c1}7"]
        lbl_cell.value = kpi["label"]
        lbl_cell.font = font_kpi_label
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_accent
        
        val_cell = ws_summary[f"{c1}8"]
        val_cell.value = kpi["val"]
        val_cell.font = font_kpi_num
        val_cell.alignment = align_center
        val_cell.number_format = kpi["fmt"]
        
        for r in [7, 8]:
            for c in range(kpi["col_start"], kpi["col_end"] + 1):
                cell = ws_summary.cell(row=r, column=c)
                cell.border = thin_border

    # Section Header for Table
    ws_summary["A10"] = "ENDPOINT PERFORMANCE SUMMARY"
    ws_summary["A10"].font = font_section
    ws_summary.row_dimensions[10].height = 20

    # Table Headers (Row 11)
    headers = [
        "Endpoint", "Total Requests", "Successful", "Failed", 
        "Success Rate", "Avg Latency (ms)", "Min Latency (ms)", 
        "Max Latency (ms)", "95% Latency (ms)", "RPS"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=11, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = thin_border
    
    ws_summary.row_dimensions[11].height = 25

    # Data Rows (Row 12+)
    row_idx = 12
    for ep, stats in endpoints.items():
        ws_summary.cell(row=row_idx, column=1, value=ep).alignment = align_left
        ws_summary.cell(row=row_idx, column=2, value=stats["total_requests"]).number_format = "#,##0"
        ws_summary.cell(row=row_idx, column=3, value=stats["successful_requests"]).number_format = "#,##0"
        ws_summary.cell(row=row_idx, column=4, value=stats["failed_requests"]).number_format = "#,##0"
        
        # Success Rate
        sr_cell = ws_summary.cell(row=row_idx, column=5, value=stats["success_rate"] / 100.0)
        sr_cell.number_format = "0.0%"
        if stats["success_rate"] >= 99.0:
            sr_cell.fill = fill_pass
            sr_cell.font = Font(name="Segoe UI", size=10, bold=True, color=PASS_TEXT_COLOR)
        else:
            sr_cell.fill = fill_fail
            sr_cell.font = Font(name="Segoe UI", size=10, bold=True, color=FAIL_TEXT_COLOR)
            
        ws_summary.cell(row=row_idx, column=6, value=stats["avg_response_time_ms"]).number_format = "0.0"
        ws_summary.cell(row=row_idx, column=7, value=stats["min_response_time_ms"]).number_format = "0.0"
        ws_summary.cell(row=row_idx, column=8, value=stats["max_response_time_ms"]).number_format = "0.0"
        ws_summary.cell(row=row_idx, column=9, value=stats["p95_response_time_ms"]).number_format = "0.0"
        ws_summary.cell(row=row_idx, column=10, value=stats["rps"]).number_format = "0.0"

        for c in range(1, 11):
            cell = ws_summary.cell(row=row_idx, column=c)
            cell.border = thin_border
            if c > 1:
                cell.alignment = align_right
                if c != 5: # Keep custom formatting for success rate
                    cell.font = font_body
            else:
                cell.font = font_body_bold
                
        ws_summary.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Total/Summary Row in table
    tot_row = row_idx
    ws_summary.cell(row=tot_row, column=1, value="Total / Average").font = font_body_bold
    ws_summary.cell(row=tot_row, column=1).alignment = align_left
    
    # Formulas for totals
    ws_summary.cell(row=tot_row, column=2, value=f"=SUM(B12:B{tot_row-1})").number_format = "#,##0"
    ws_summary.cell(row=tot_row, column=3, value=f"=SUM(C12:C{tot_row-1})").number_format = "#,##0"
    ws_summary.cell(row=tot_row, column=4, value=f"=SUM(D12:D{tot_row-1})").number_format = "#,##0"
    
    # Success Rate formula
    sr_formula = ws_summary.cell(row=tot_row, column=5, value=f"=C{tot_row}/B{tot_row}")
    sr_formula.number_format = "0.0%"
    sr_formula.font = font_body_bold
    
    # Averages
    ws_summary.cell(row=tot_row, column=6, value=f"=AVERAGE(F12:F{tot_row-1})").number_format = "0.0"
    ws_summary.cell(row=tot_row, column=7, value=f"=MIN(G12:G{tot_row-1})").number_format = "0.0"
    ws_summary.cell(row=tot_row, column=8, value=f"=MAX(H12:H{tot_row-1})").number_format = "0.0"
    ws_summary.cell(row=tot_row, column=9, value=f"=AVERAGE(I12:I{tot_row-1})").number_format = "0.0"
    ws_summary.cell(row=tot_row, column=10, value=f"=SUM(J12:J{tot_row-1})").number_format = "0.0"

    for c in range(1, 11):
        cell = ws_summary.cell(row=tot_row, column=c)
        cell.border = double_bottom_border
        cell.font = font_body_bold
        if c > 1:
            cell.alignment = align_right
            
    ws_summary.row_dimensions[tot_row].height = 22

    # Add a visual chart showing Response Times
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Average Latency by Endpoint"
    chart.y_axis.title = "Response Time (ms)"
    chart.x_axis.title = "Endpoint"
    
    # Reference for values (Avg Latency is Column F, row 11 to tot_row-1)
    data_ref = Reference(ws_summary, min_col=6, min_row=11, max_row=tot_row-1)
    # Reference for categories (Endpoint is Column A, row 12 to tot_row-1)
    cats_ref = Reference(ws_summary, min_col=1, min_row=12, max_row=tot_row-1)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None # No legend needed for single series
    
    # Position chart below the table
    ws_summary.add_chart(chart, f"A{tot_row+3}")
    chart.width = 16
    chart.height = 10

    # ------------------ SHEET 2: RAW DATA SAMPLES ------------------
    # Write a secondary sheet containing the raw request log
    raw_results = data.get("raw_results", [])
    if raw_results:
        ws_raw = wb.create_sheet(title="Response Logs")
        ws_raw.views.sheetView[0].showGridLines = True
        
        # Raw headers
        raw_headers = ["Timestamp (Relative s)", "Endpoint", "Status Code", "Latency (ms)", "Status"]
        for col_idx, h in enumerate(raw_headers, 1):
            cell = ws_raw.cell(row=1, column=col_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = align_center
            cell.border = thin_border
            
        ws_raw.row_dimensions[1].height = 25
        
        # Sort raw results by timestamp
        raw_results_sorted = sorted(raw_results, key=lambda x: x["timestamp"])
        test_start_t = raw_results_sorted[0]["timestamp"] if raw_results_sorted else 0
        
        # To avoid creating massive Excel files, write up to 5000 records
        limit = min(len(raw_results_sorted), 5000)
        print(f"Writing {limit} sample logs to 'Response Logs' sheet...")
        
        for idx, r in enumerate(raw_results_sorted[:limit], 2):
            rel_t = r["timestamp"] - test_start_t
            ws_raw.cell(row=idx, column=1, value=rel_t).number_format = "0.000"
            ws_raw.cell(row=idx, column=2, value=r["endpoint"]).alignment = align_left
            ws_raw.cell(row=idx, column=3, value=r["status_code"]).number_format = "0"
            ws_raw.cell(row=idx, column=4, value=r["latency_ms"]).number_format = "0.0"
            
            status_cell = ws_raw.cell(row=idx, column=5, value="SUCCESS" if r["success"] else "FAIL")
            status_cell.alignment = align_center
            if r["success"]:
                status_cell.font = Font(name="Segoe UI", size=9, bold=True, color=PASS_TEXT_COLOR)
                status_cell.fill = fill_pass
            else:
                status_cell.font = Font(name="Segoe UI", size=9, bold=True, color=FAIL_TEXT_COLOR)
                status_cell.fill = fill_fail

            for c in range(1, 6):
                cell = ws_raw.cell(row=idx, column=c)
                cell.border = thin_border
                cell.font = font_body
                if c in [1, 3, 4]:
                    cell.alignment = align_right
                    
            if idx % 1000 == 0:
                print(f"  Wrote {idx} log rows...")
                
        # Auto-fit columns for both sheets
        for sheet in [ws_summary, ws_raw]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Don't size columns A-H by the merged title or charts
                    if cell.row > 2 and cell.value and not isinstance(cell.value, str) and not cell.coordinate in sheet.merged_cells:
                        val_str = str(cell.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    elif cell.row > 2 and cell.value and isinstance(cell.value, str):
                        # Simple length check for string
                        val_str = cell.value
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 35)
    else:
        # Auto-fit Summary sheet only
        for col in ws_summary.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 2 and cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws_summary.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 35)

    # Specific layouts adjustments
    ws_summary.column_dimensions["A"].width = 25 # Endpoint name

    # Save report
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel Load Test Report successfully generated at: {os.path.abspath(output_path)}")

if __name__ == "__main__":
    main()
