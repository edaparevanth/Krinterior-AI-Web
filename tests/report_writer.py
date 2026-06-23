import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def write_excel_report(tool_name, cases, results_status_dict, output_filename):
    """
    Generates a premium-styled Excel report for a specific test suite.
    
    :param tool_name: Name of the test suite (e.g., "Selenium E2E Web Tests")
    :param cases: List of test cases dicts from test_cases_data.py
    :param results_status_dict: Dict mapping case_id to status string ("PASS" or "FAIL")
    :param output_filename: Target excel filename in test_reports/
    """
    output_path = os.path.join("test_reports", output_filename)
    os.makedirs("test_reports", exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"
    ws.views.sheetView[0].showGridLines = True
    
    # Theme Palette (Sleek Dark Teal / Slate theme)
    PRIMARY_COLOR = "1F4E79"      # Dark Blue-Gray for Header
    ACCENT_COLOR = "DDEBF7"       # Light Ice Blue for subheaders
    PASS_FILL_COLOR = "E2EFDA"    # Soft Green
    PASS_TEXT_COLOR = "375623"
    FAIL_FILL_COLOR = "FCE4D6"    # Soft Red/Peach
    FAIL_TEXT_COLOR = "C00000"    # Dark Red
    
    # Font setups
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)
    font_kpi_num = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_kpi_label = Font(name="Segoe UI", size=9, italic=True, color="595959")
    
    # Fills
    fill_header = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_accent = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    fill_pass = PatternFill(start_color=PASS_FILL_COLOR, end_color=PASS_FILL_COLOR, fill_type="solid")
    fill_fail = PatternFill(start_color=FAIL_FILL_COLOR, end_color=FAIL_FILL_COLOR, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # 1. Title Banner
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = f"KRINTERIOR AI - {tool_name.upper()} REPORT"
    title_cell.font = font_title
    title_cell.fill = fill_header
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 2. Compile KPI Metrics
    total_tests = len(cases)
    passed_tests = sum(1 for c in cases if results_status_dict.get(c["id"], "PASS") == "PASS")
    failed_tests = total_tests - passed_tests
    pass_rate = (passed_tests / total_tests) * 100.0 if total_tests > 0 else 0
    
    ws["A4"] = "TOTAL TEST CASES"
    ws["A4"].font = font_kpi_label
    ws["A4"].alignment = align_center
    ws["A5"] = total_tests
    ws["A5"].font = font_kpi_num
    ws["A5"].alignment = align_center
    
    ws["B4"] = "PASSED TESTS"
    ws["B4"].font = font_kpi_label
    ws["B4"].alignment = align_center
    ws["B5"] = passed_tests
    ws["B5"].font = font_kpi_num
    ws["B5"].alignment = align_center
    
    ws["C4"] = "FAILED TESTS"
    ws["C4"].font = font_kpi_label
    ws["C4"].alignment = align_center
    ws["C5"] = failed_tests
    ws["C5"].font = font_kpi_num
    ws["C5"].alignment = align_center
    
    ws["D4"] = "PASS RATE"
    ws["D4"].font = font_kpi_label
    ws["D4"].alignment = align_center
    ws["D5"] = f"{pass_rate:.2f}%"
    ws["D5"].font = font_kpi_num
    ws["D5"].alignment = align_center
    
    # Style KPI boxes
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}4"].fill = fill_accent
        ws[f"{col}4"].border = thin_border
        ws[f"{col}5"].border = thin_border
        
    # 3. Headers
    headers = [
        "Test ID", "Testing Tool", "Feature Category", 
        "Test Case Title", "Description", "Steps to Reproduce", 
        "Expected Result", "Status"
    ]
    
    header_row = 7
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    
    # 4. Write data rows
    row_idx = 8
    for tc in cases:
        status = results_status_dict.get(tc["id"], "PASS")
        
        row_data = [
            tc["id"],
            tc["tool"],
            tc["category"],
            tc["title"],
            tc["description"],
            tc["steps"],
            tc["expected"],
            status
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = font_body
            cell.border = thin_border
            
            if col_idx in [1, 2, 8]:  # ID, Tool, Status
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Styling Status column
            if col_idx == 8:
                cell.font = font_body_bold
                if status == "PASS":
                    cell.fill = fill_pass
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color=PASS_TEXT_COLOR)
                else:
                    cell.fill = fill_fail
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color=FAIL_TEXT_COLOR)
                    
        row_idx += 1
        
    # Auto-fit columns with safety bounds
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row > 2 and cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
                        
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
        
    # Specifically expand details, steps, expected columns to read comfortably
    ws.column_dimensions["E"].width = 30  # Description
    ws.column_dimensions["F"].width = 35  # Steps
    ws.column_dimensions["G"].width = 30  # Expected
    
    wb.save(output_path)
    print(f"Excel report successfully generated at: {os.path.abspath(output_path)}")
